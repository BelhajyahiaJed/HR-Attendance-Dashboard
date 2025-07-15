import streamlit as st
import pandas as pd
import plotly.express as px
import io
import os 
from pandas import Timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import numpy as np
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader
import copy

# --- Load config from Streamlit secrets ---
try:
    config = yaml.load(st.secrets["auth_config"]["yaml"], Loader=SafeLoader)
except Exception as e:
    st.error("Failed to load authentication config from secrets.")
    st.stop()

# --- Setup authenticator ---
authenticator = stauth.Authenticate(
    credentials=config["credentials"],
    cookie_name=config["cookie"]["name"],
    cookie_key=config["cookie"]["key"],
    cookie_expiry_days=config["cookie"]["expiry_days"]
)

# --- Perform login ---
authenticator.login(location='main')

authentication_status = st.session_state.get("authentication_status")
name = st.session_state.get("name")
username = st.session_state.get("username")

if authentication_status is False:
    st.error("Incorrect username or password.")
    st.stop()
elif authentication_status is None:
    st.warning("Please log in to access the dashboard.")
    st.stop()

# User is authenticated
authenticator.logout("Logout", "sidebar")
st.sidebar.success(f"Welcome {name} 👋")
st.title("HR Attendance Dashboard")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Root variables for dark theme */
    :root {
        --primary-color: #6366f1;
        --secondary-color: #8b5cf6;
        --accent-color: #06b6d4;
        --success-color: #10b981;
        --warning-color: #f59e0b;
        --error-color: #ef4444;
        --dark-bg: #0f172a;
        --card-bg: #1e293b;
        --sidebar-bg: #334155;
        --text-primary: #f1f5f9;
        --text-secondary: #cbd5e1;
        --text-accent: #e2e8f0;
        --border-color: #475569;
        --hover-bg: #2d3748;
        --shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.3), 0 4px 6px -2px rgba(0, 0, 0, 0.1);
        --gradient-primary: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
        --gradient-secondary: linear-gradient(135deg, #06b6d4 0%, #3b82f6 100%);
        --gradient-success: linear-gradient(135deg, #10b981 0%, #059669 100%);
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* Main app container with dark background */
    .main {
        background: var(--dark-bg);
        min-height: 100vh;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        color: var(--text-primary);
    }
    
    /* Content wrapper - dark card style */
    .block-container {
        background: var(--card-bg);
        border-radius: 20px;
        padding: 2rem;
        box-shadow: var(--shadow);
        margin: 1rem auto;
        max-width: 1200px;
        border: 1px solid var(--border-color);
    }
    
    /* Title styling with gradient text */
    h1 {
        background: var(--gradient-primary);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-weight: 700;
        font-size: 3rem;
        text-align: center;
        margin-bottom: 0.5rem;
        letter-spacing: -0.02em;
    }
    
    /* Subtitle styling */
    .stApp > div > div > div > div > div:nth-child(2) {
        text-align: center;
        color: var(--text-secondary);
        font-size: 1.1rem;
        margin-bottom: 2rem;
    }
    
    /* Subheaders with better contrast */
    h2, h3 {
        color: var(--text-primary) !important;
        font-weight: 600;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    
    /* Sidebar styling with dark theme */
    .css-1d391kg, .st-emotion-cache-6qob1r, .st-emotion-cache-1cypcdb {
        background: var(--sidebar-bg) !important;
        border-radius: 15px;
        margin: 1rem;
        padding: 1.5rem;
        border: 1px solid var(--border-color);
    }
    
    /* Sidebar header with high contrast */
    .css-1d391kg h2, .st-emotion-cache-6qob1r h2, .st-emotion-cache-1cypcdb h2 {
        color: var(--text-accent) !important;
        font-weight: 700;
        font-size: 1.2rem;
        margin-bottom: 1.5rem;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
    }
    
    /* Sidebar text with high contrast */
    .css-1d391kg .stMarkdown, .st-emotion-cache-6qob1r .stMarkdown, .st-emotion-cache-1cypcdb .stMarkdown {
        color: var(--text-accent) !important;
    }
    
    .css-1d391kg p, .st-emotion-cache-6qob1r p, .st-emotion-cache-1cypcdb p {
        color: var(--text-accent) !important;
        font-weight: 500;
    }
    
    /* Sidebar buttons with better styling */
    .stButton > button {
        background: linear-gradient(135deg, var(--primary-color) 0%, var(--secondary-color) 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 0.95rem;
        transition: all 0.3s ease;
        width: 100%;
        margin-bottom: 0.5rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(99, 102, 241, 0.3);
        background: linear-gradient(135deg, var(--secondary-color) 0%, var(--primary-color) 100%) !important;
    }
    
    /* File uploader with dark theme */
    .stFileUploader {
        background: var(--hover-bg);
        border: 2px dashed var(--primary-color);
        border-radius: 15px;
        padding: 2rem;
        text-align: center;
        transition: all 0.3s ease;
        margin: 1rem 0;
    }
    
    .stFileUploader:hover {
        border-color: var(--secondary-color);
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.1) 0%, rgba(139, 92, 246, 0.1) 100%);
    }
    
    .stFileUploader label {
        color: var(--text-primary) !important;
        font-weight: 500;
    }
    
    /* Upload button */
    .stFileUploader button {
        background: var(--gradient-primary) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stFileUploader button:hover {
        transform: translateY(-2px);
        box-shadow: 0 5px 15px rgba(99, 102, 241, 0.3);
    }
    
    /* Dataframes with dark theme */
    .stDataFrame {
        background: var(--hover-bg);
        border-radius: 15px;
        overflow: hidden;
        box-shadow: var(--shadow);
        margin: 1rem 0;
        border: 1px solid var(--border-color);
    }
    
    .stDataFrame table {
        background: var(--hover-bg) !important;
        color: var(--text-primary) !important;
    }
    
    .stDataFrame th {
        background: var(--primary-color) !important;
        color: white !important;
        font-weight: 600;
    }
    
    .stDataFrame td {
        background: var(--hover-bg) !important;
        color: var(--text-primary) !important;
        border-bottom: 1px solid var(--border-color);
    }
    
    .stDataFrame tr:hover td {
        background: var(--card-bg) !important;
    }
    
    /* Download button */
    .stDownloadButton > button {
        background: var(--gradient-success) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 0.95rem;
        transition: all 0.3s ease;
        margin: 1rem 0;
    }
    
    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 5px 15px rgba(16, 185, 129, 0.3);
    }
    
    /* Selectbox with dark theme */
    .stSelectbox > div > div {
        background: var(--hover-bg) !important;
        border: 2px solid var(--border-color) !important;
        border-radius: 10px;
        transition: all 0.3s ease;
    }
    
    .stSelectbox > div > div:hover {
        border-color: var(--primary-color) !important;
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.2);
    }
    
    .stSelectbox label {
        color: var(--text-primary) !important;
        font-weight: 500;
    }
    
    .stSelectbox div[data-baseweb="select"] > div {
        background: var(--hover-bg) !important;
        color: var(--text-primary) !important;
    }
    
    /* Charts with dark theme */
    .stPlotlyChart {
        background: var(--hover-bg);
        border-radius: 15px;
        padding: 1rem;
        box-shadow: var(--shadow);
        margin: 1rem 0;
        border: 1px solid var(--border-color);
    }
    
    /* Info boxes with dark theme */
    .stInfo {
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.2) 0%, rgba(139, 92, 246, 0.2) 100%);
        border-left: 4px solid var(--primary-color);
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
    }
    
    .stInfo div {
        color: var(--text-primary) !important;
    }
    
    .stError {
        background: linear-gradient(135deg, rgba(239, 68, 68, 0.2) 0%, rgba(239, 68, 68, 0.1) 100%);
        border-left: 4px solid var(--error-color);
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
    }
    
    .stError div {
        color: var(--text-primary) !important;
    }
    
    .stSuccess {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.2) 0%, rgba(5, 150, 105, 0.2) 100%);
        border-left: 4px solid var(--success-color);
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
    }
    
    .stSuccess div {
        color: var(--text-primary) !important;
    }
    
    /* Text styling with proper contrast */
    .stMarkdown {
        color: var(--text-primary) !important;
        line-height: 1.6;
    }
    
    .stMarkdown p, .stMarkdown div {
        color: var(--text-primary) !important;
    }
    
    .stMarkdown strong {
        color: var(--text-accent) !important;
        font-weight: 600;
    }
    
    /* Dividers with gradient */
    hr {
        border: none;
        height: 2px;
        background: var(--gradient-primary);
        margin: 2rem 0;
        border-radius: 1px;
    }
    
    /* Metrics styling */
    .metric-card {
        background: var(--hover-bg);
        border-radius: 15px;
        padding: 1.5rem;
        box-shadow: var(--shadow);
        border: 1px solid var(--border-color);
        margin: 1rem 0;
        transition: all 0.3s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.2), 0 10px 10px -5px rgba(0, 0, 0, 0.1);
        border-color: var(--primary-color);
    }
    
    /* Animation for loading */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .main > div {
        animation: fadeIn 0.6s ease-out;
    }
    
    /* Responsive design */
    @media (max-width: 768px) {
        .block-container {
            padding: 1rem;
            margin: 0.5rem;
        }
        
        h1 {
            font-size: 2rem;
        }
        
        .stButton > button {
            padding: 0.5rem 1rem;
            font-size: 0.9rem;
        }
    }
    
    /* Custom scrollbar with dark theme */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: var(--card-bg);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: var(--gradient-primary);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: var(--secondary-color);
    }
    
    /* Ensure all text in main content is properly colored */
    .stApp .main .block-container * {
        color: var(--text-primary);
    }
    
    /* Force sidebar text to be visible */
    .css-1d391kg *, .st-emotion-cache-6qob1r *, .st-emotion-cache-1cypcdb * {
        color: var(--text-accent) !important;
    }
    
    /* Navigation text specifically */
    .css-1d391kg .stButton label, .st-emotion-cache-6qob1r .stButton label, .st-emotion-cache-1cypcdb .stButton label {
        color: white !important;
        font-weight: 600;
    }
    
    /* Fix for sidebar dividers */
    .css-1d391kg hr, .st-emotion-cache-6qob1r hr, .st-emotion-cache-1cypcdb hr {
        background: var(--text-accent) !important;
        opacity: 0.3;
    }
    
    /* Additional fixes for Streamlit components */
    .stTextInput > div > div > input {
        background: var(--hover-bg) !important;
        color: var(--text-primary) !important;
        border: 2px solid var(--border-color) !important;
    }
    
    .stTextInput > div > div > input:focus {
        border-color: var(--primary-color) !important;
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.2);
    }
    
    .stTextInput label {
        color: var(--text-primary) !important;
    }
    
    /* Fix for select dropdown options */
    .stSelectbox div[data-baseweb="select"] div[role="listbox"] {
        background: var(--hover-bg) !important;
        border: 1px solid var(--border-color) !important;
    }
    
    .stSelectbox div[data-baseweb="select"] div[role="option"] {
        background: var(--hover-bg) !important;
        color: var(--text-primary) !important;
    }
    
    .stSelectbox div[data-baseweb="select"] div[role="option"]:hover {
        background: var(--card-bg) !important;
    }
</style>
""", unsafe_allow_html=True)

# Function to process time formats (handle strings, Timedelta, and numbers)
def clean_time_format(value):
    try:
        if isinstance(value, Timedelta):
            return value.total_seconds() / 3600.0
        elif isinstance(value, str) and ":" in value:
            return float(value.replace(":", "."))
        elif pd.isna(value):
            return 0.0
        else:
            return float(value)
    except (ValueError, TypeError):
        return 0.0

# Function to process the uploaded Excel file
def process_excel_file(file):
    try:
        # Read Excel file
        df = pd.read_excel(file, sheet_name=0)
        
        # Clean column names
        df.columns = [col.strip() for col in df.columns]
        
        # Apply time format cleaning to relevant columns (D and H)
        if len(df.columns) >= 8:  # Ensure columns D (index 3) and H (index 7) exist
            df.iloc[:, 3] = df.iloc[:, 3].apply(clean_time_format)  # Heures Normales
            df.iloc[:, 7] = df.iloc[:, 7].apply(clean_time_format)  # Congé annuel
        
        # Initialize Presence column
        df['Presence'] = 0.0
        
        # Process workers based on Jours de Présences (column L, index 11)
        for idx, row in df.iterrows():
            jours_presence = row.iloc[11] if len(df.columns) > 11 else None
            if pd.isna(jours_presence) or jours_presence == '':
                # Hourly worker: Apply formula D - ((H * 8) + 14)
                heures_normales = float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0.0
                conge_annuel = float(row.iloc[7]) if pd.notna(row.iloc[7]) else 0.0
                df.at[idx, 'Presence'] = heures_normales - ((conge_annuel * 8) + 14)
            else:
                # Daily worker: Use Jours de Présences directly
                df.at[idx, 'Presence'] = float(jours_presence) if pd.notna(jours_presence) else 0.0
        
        # Convert Service en cours to string and handle NaN
        if 'Service en cours' in df.columns:
            df['Service en cours'] = df['Service en cours'].fillna('Unknown').astype(str)
        
        # Create output DataFrame with required columns
        output_df = pd.DataFrame({
            'Matr': df['Matricule'],
            'Nom & Prénom': df['Nom'] + ' ' + df['Prénom'],
            'Présence': df['Presence'],
            'H Supp 75%': df.get('H SUP 75% Hebdomadaire 24/5-23/6', ''),
            'Congé': df.get('Congé annuel 24/5-23/6', ''),
            'Congé spécial': '',
            'Feries': df.get('Nombre de jours fériés 24/5-23/6', ''),
            'Chom Tech': df.get('Chomage technique 24/5-23/6', ''),
            'Prime de Rendement': '',
            'Avance sur salaire': '',
            'Prêt /Salaire': '',
            'Observations': 'Prime anciennete:30 dt',
            'Service en cours': df.get('Service en cours', 'Unknown')
        })
        
        return df, output_df
    except Exception as e:
        st.error(f"Failed to process Excel file: {str(e)}")
        return None, None

# Function to create styled Excel file
def create_styled_excel(df, output_buffer):
    wb = Workbook()
    ws = wb.active
    ws.title = "Feuil1"
    
    # Define font
    calibri_font = Font(name='Calibri', size=11, bold=False)
    calibri_bold = Font(name='Calibri', size=11, bold=True)
    
    # Add title, subtitle, and reference line
    ws['A1'] = "PRINCE MEDICAL INDUSTRY PMI SARL"
    ws.merge_cells('A1:L1')
    ws['A1'].font = calibri_bold
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = "ETAT DE POINTAGE Juin 2025"
    ws.merge_cells('A2:L2')
    ws['A2'].font = calibri_bold
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A3'] = "Présence 192 heures"
    ws.merge_cells('A3:L3')
    ws['A3'].font = calibri_font
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.append([])  # Blank row
    
    # Define header row
    headers = ["Matr", "Nom & Prénom", "Présence", "H Supp 75%", "Congé", "Congé spécial",
            "Feries", "Chom Tech", "Prime de Rendement", "Avance sur salaire", "Prêt /Salaire", "Observations"]
    
    # Group employees into chunks of 20
    chunk_size = 20
    row_idx = 5
    table_count = 0
    
    for i in range(0, len(df), chunk_size):
        # Add header
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = calibri_bold
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        chunk = df.iloc[i:i + chunk_size]
        for _, row in chunk.iterrows():
            row_idx += 1
            ws.append([row[col] for col in headers])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).font = calibri_font
        
        # Create table
        table_ref = f"A{row_idx - len(chunk)}:L{row_idx}"
        table = Table(displayName=f"Table{table_count}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(table)
        table_count += 1
        
        # Add visa lines
        row_idx += 1
        ws.append([])  # Blank row
        row_idx += 1
        ws.append(["Visa RAP"] + [""] * 8 + ["Visa Direction Général"])
        ws.cell(row=row_idx, column=1).font = calibri_font
        ws.cell(row=row_idx, column=10).font = calibri_font
        row_idx += 1
        ws.append([])  # Blank row
        row_idx += 1
    
    # Adjust column widths based on data rows (starting from row 5)
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(5, ws.max_row + 1):  # Start from header row
            cell = ws.cell(row=row, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_buffer)

# Streamlit app
st.title("HR Attendance Dashboard")
st.write("Navigate using the sidebar to view processed data or statistics.")

# Initialize session state
if 'output_df' not in st.session_state:
    st.session_state.output_df = None
if 'original_df' not in st.session_state:
    st.session_state.original_df = None
if 'page' not in st.session_state:
    st.session_state.page = "Upload"

# Sidebar for navigation with buttons
st.sidebar.header("Navigation")
if st.session_state.page != "Upload":
    if st.sidebar.button("Data Processing"):
        st.session_state.page = "Data Processing"
        st.rerun()
    if st.sidebar.button("Stats"):
        st.session_state.page = "Stats"
        st.rerun()
else:
    st.sidebar.write("Upload a file to enable navigation.")

# Handle file upload and page navigation
if st.session_state.page == "Upload" and st.session_state.output_df is None:
    # Center the file uploader with wider span
    col1, col2, col3 = st.columns([1, 4, 1])
    with col2:
        uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])
    
    if uploaded_file is not None:
        original_df, output_df = process_excel_file(uploaded_file)
        if output_df is not None:
            st.session_state.original_df = original_df
            st.session_state.output_df = output_df
            st.session_state.page = "Data Processing"
            st.rerun()  # Redirect to Data Processing page
else:
    # Service filter for Stats page only
    if st.session_state.output_df is not None:
        try:
            services = sorted([str(x) for x in st.session_state.output_df['Service en cours'].unique() if pd.notna(x)])
            services = ["All Services"] + services  # Add All Services option
        except Exception as e:
            st.error(f"Error sorting services: {str(e)}")
            services = []
        
        if st.session_state.page == "Data Processing":
            filtered_df = st.session_state.output_df  # No filtering
        else:  # Stats page
            selected_service = st.selectbox(
                "Filter by Service", 
                services,
                index=0  # Default to All Services
            )
            if selected_service == "All Services":
                filtered_df = st.session_state.output_df
            else:
                filtered_df = st.session_state.output_df[
                    st.session_state.output_df['Service en cours'] == selected_service
                ]
    else:
        filtered_df = None

    # Page logic
    if st.session_state.page == "Data Processing" and st.session_state.output_df is not None:
        # Display processed data
        st.subheader("Processed Data")
        st.dataframe(filtered_df[['Matr', 'Nom & Prénom', 'Présence', 'H Supp 75%', 
                                'Congé', 'Feries', 'Chom Tech', 'Observations']])
        
        # Download processed data
        output = io.BytesIO()
        create_styled_excel(filtered_df, output)
        st.download_button(
            label="Download Processed Data",
            data=output.getvalue(),
            file_name="etat_de_pointage.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    elif st.session_state.page == "Stats" and st.session_state.output_df is not None:
        st.subheader("Attendance Statistics")
        
        # Summary statistics
        st.write("**Summary Statistics for Presence**")
        stats = filtered_df['Présence'].describe()
        st.write(stats)
        
        # Anomaly detection (using z-scores)
        st.write("**Anomaly Detection**")
        worker_type = st.session_state.original_df['Jours de Présences 24/5-23/6'].apply(
            lambda x: 'Hourly' if pd.isna(x) or x == '' else 'Daily'
        )
        filtered_worker_type = worker_type[st.session_state.output_df['Service en cours'].isin(
            [selected_service] if selected_service != "All Services" else services[1:]
        )]
        hourly_workers = filtered_df[filtered_worker_type == 'Hourly']
        daily_workers = filtered_df[filtered_worker_type == 'Daily']
        
        # Top 5 and Worst 5 for Hourly Workers
        if not hourly_workers.empty:
            hourly_workers = hourly_workers.copy()
            hourly_workers['Z-Score'] = np.abs((hourly_workers['Présence'] - hourly_workers['Présence'].mean()) / 
                                            hourly_workers['Présence'].std())
            st.write("**Hourly Workers**")
            
            # Top 5 (highest presence)
            top_5_hourly = hourly_workers.nlargest(5, 'Présence')[['Matr', 'Nom & Prénom', 'Présence', 'Service en cours']]
            st.write("Top 5 Hourly Workers (Highest Presence)")
            st.dataframe(top_5_hourly)
            
            # Worst 5 (lowest presence)
            worst_5_hourly = hourly_workers.nsmallest(5, 'Présence')[['Matr', 'Nom & Prénom', 'Présence', 'Service en cours']]
            st.write("Worst 5 Hourly Workers (Lowest Presence)")
            st.dataframe(worst_5_hourly)
        
        # Top 5 and Worst 5 for Daily Workers
        if not daily_workers.empty:
            daily_workers = daily_workers.copy()
            daily_workers['Z-Score'] = np.abs((daily_workers['Présence'] - daily_workers['Présence'].mean()) / 
                                            daily_workers['Présence'].std())
            st.write("**Daily Workers**")
            
            # Top 5 (highest presence)
            top_5_daily = daily_workers.nlargest(5, 'Présence')[['Matr', 'Nom & Prénom', 'Présence', 'Service en cours']]
            st.write("Top 5 Daily Workers (Highest Presence)")
            st.dataframe(top_5_daily)
            
            # Worst 5 (lowest presence)
            worst_5_daily = daily_workers.nsmallest(5, 'Présence')[['Matr', 'Nom & Prénom', 'Présence', 'Service en cours']]
            st.write("Worst 5 Daily Workers (Lowest Presence)")
            st.dataframe(worst_5_daily)
        
        # Visualizations
        st.subheader("Visualizations")
        
        # Bar chart: Average Presence by Service
        worker_type = st.session_state.original_df['Jours de Présences 24/5-23/6'].apply(
            lambda x: 'Hourly' if pd.isna(x) or x == '' else 'Daily'
        )
        df_with_worker_type = filtered_df.copy()
        df_with_worker_type['Worker Type'] = worker_type
        avg_presence_by_service = df_with_worker_type.groupby(['Service en cours', 'Worker Type'])['Présence'].mean().reset_index()
        fig_bar = px.bar(
            avg_presence_by_service,
            x='Service en cours',
            y='Présence',
            color='Worker Type',
            title="Average Presence by Service (Hours/Days)",
            labels={'Présence': 'Average Presence (Hours/Days)', 'Service en cours': 'Service'},
            barmode='group',
            color_discrete_sequence=['#1f77b4', '#ff7f0e'],
            range_y=[0, 300]  # Cap y-axis at 300
        )
        st.plotly_chart(fig_bar)
        
        # Box plot: Presence Distribution by Service
        fig_box = px.box(
            filtered_df,
            x='Service en cours',
            y='Présence',
            title="Presence Distribution by Service",
            labels={'Présence': 'Presence (Hours/Days)', 'Service en cours': 'Service'},
            hover_data=['Nom & Prénom'],
            range_y=[0, 300]  # Cap y-axis at 300
        )
        st.plotly_chart(fig_box)
        
        # Histogram of Presence
        st.write("**Presence Distribution**")
        fig_hist = px.histogram(
            filtered_df,
            x='Présence',
            color=filtered_worker_type,
            title="Distribution of Presence (Hourly vs Daily Workers)",
            labels={'Présence': 'Presence (Hours/Days)', 'color': 'Worker Type'},
            nbins=20
        )
        st.plotly_chart(fig_hist)
    
    else:
        st.info("Please upload an Excel file to begin.")
